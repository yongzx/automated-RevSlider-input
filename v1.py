import openpyxl #open Excel file
import copy #for deep copy
import re

#to access individual Topic data, ExcelData_instance[row_number-1][column_number][category]

#category --> title, id, start time, end time
TITLE, ID, START_TIME, END_TIME = range(4)
TEXT_INFO = 3

#ExcelData contains all the data in a Excel sheet
class ExcelData:
    def __init__(self):
        self.m_list = []

    def __getitem__(self,index):
        return self.m_list[index]

    def __len__(self):
        return len(self.m_list)

    def __str__(self):
        return str(self.m_list)

    def __repr__(self):
        return self.__str__()

    #add the TopicData instance containing video infomration of a particular topic into ExcelData m_list
    def read(self,topicdata,worksheet):
        max_row = worksheet.max_row
        max_column = worksheet.max_column
        #print(max_row, max_column)
        for row in range(2,max_row+1):
            for col in range(4,max_column+1):
                cell_var = worksheet.cell(row = row, column = col).value    #cell_val is a string
                #print(cell_var)
                if cell_var is None:
                    continue
                list1 = cell_var.split('\n')
                topicdata.add(col, list1)
            topicdata.splitTime()                         #split the duration into start time and end time
            topic_deepcopy = copy.deepcopy(topicdata)   #deepcopy to avoid aliasing
            self.m_list.append(topic_deepcopy)
            topicdata.empty()


# TopicData object contains all the video information for a particular topic
class TopicData:
    
    def __init__(self):
        self.m_dict = {}

    def __getitem__(self,key):
        return self.m_dict[key]

    def __len__(self):
        return len(self.m_dict)

    def __str__(self):
        return str(self.m_dict)
    
    def __repr__(self):
        return self.__str__()

    #implement deepcopy to avoid aliasing when the TopicData instance is appended into VideoData instance 
    def __deepcopy__(self, memo):
        return copy.deepcopy(self.m_dict)

    #add the video information of one cell (in list form) into m_dict. 
    #col_n -- column number
    #list_var contains 3 things of each column: title, iD, raw time
    def add(self, col_n, list_var):
        self.m_dict[col_n - TEXT_INFO] = list_var

    #empty the m_dict
    def empty(self):
        self.m_dict = {}


    #split the raw time into start time and end time
    def splitTime(self):
        
        for i in range(1, len(self.m_dict)+1):

            split_time_list = []
            split_time_list.append(self.m_dict[i][TITLE])
            split_time_list.append(self.m_dict[i][ID])

            for raw_time_input in range(2, len(self.m_dict[i])):
                
                temp_str = self.m_dict[i][raw_time_input]
                #print(temp_str)
                #if the raw time is 'whole' --> split into '' and ''
                if temp_str == 'whole':
                    split_time_list.append('')
                    split_time_list.append('')

                #the raw time is split into start time and end time     
                else:
                    #assuming youtube ID doesn't have hyphen
                    temp_list = temp_str.split('-')        #list temp_list contains 2 items: start time, end time
                    if len(temp_list) == 1:
                        split_time_list.append(temp_list[0])
                    else:    
                        for j in range(0,len(temp_list)):
                            if temp_list[j] == 'start' or temp_list[j] == 'end':
                                temp_list[j] = ''
                        split_time_list.append(temp_list[0])
                        split_time_list.append(temp_list[1])
            self.m_dict[i]=split_time_list

video1 = ExcelData()
topic1 = TopicData()


#read in Excel file and sheet name
wb1 = openpyxl.load_workbook("test.xlsx")
sheet1 = wb1.get_sheet_by_name("Sheet1")
video1.read(topic1,sheet1)
#print(video1)
#workbook_name = input("Excel doc: ")
#sheet_name = input("Sheet name: ")
#print(len(video1))
#print('s:5:"title";s:2:"i1";s:5:"alias";s:2:"i1";s:9:"shortcode";s:23:"[rev_slider alias="i1"]"')
#print(r's:5:"title";s:2:"i1";s:5:"alias";s:2:"i1";s:9:"shortcode";s:23:"[rev_slider alias="i1"]"')


class Output:
    """
    read template files.
    replace the information in the template file with the data from the ExcelData.
    """
    def __init__(self, head_file, body_file, end_file):
        tmp_head_file = open(head_file, 'r+')
        tmp_body_file = open(body_file, 'r+')
        tmp_end_file = open(end_file,'r+')

        self.head_text0 = tmp_head_file.read()
        self.body_text0 = tmp_body_file.read()
        self.end_text0 = tmp_end_file.read()
        
        self.head_text = copy.deepcopy(self.head_text0)
        self.body_text = copy.deepcopy(self.body_text0)

        self.list_body_text = []
        
    def modify_HeadSegment(self,n):
        #segment number
        #row_number = range(1,len(video1))
        string_to_replace = r's:5:"title";s:2:"i1";s:5:"alias";s:2:"i1";s:9:"shortcode";s:23:"\[rev_slider alias="i1"]"'
        replacement_in_tuple = ('s:5:"title";s:2:"i',str(n+1),'";s:5:"alias";s:2:"i',str(n+1),\
            '";s:9:"shortcode";s:23:"[rev_slider alias="i', str(n+1), '"]"')
        replacement_string = ''.join(replacement_in_tuple)
        changed = re.sub(string_to_replace, replacement_string, self.head_text)
        self.head_text = changed

    def modify_HeadSlidesN(self, n):
        number_of_slides = len(video1[n])
        string_to_replace = r's:17:"custom_javascript";s:0:"";}s:6:"slides";a:7:{'
        replacement_in_tuple = ('s:17:"custom_javascript";s:0:"";}s:6:"slides";a:', str(number_of_slides), ':{')
        replacement_string = ''.join(replacement_in_tuple)
        changed = re.sub(string_to_replace, replacement_string, self.head_text)
        self.head_text = changed

    def modify_Body(self, row, txtfile1, txtfile2):
        for key in range(0,len(video1[row])):
            if key == 3 or key == 4 or key == 5 or key == 8 or key == 11:
                self.modify_Body1(row, key, txtfile1)
            else:
                self.modify_Body2(row, key, txtfile2) 

    def modify_Body1(self, row, key, txtfile):
        index_to_replace = r'xi:0';
        index_replacement_in_tuple = ('i:', str(key))
        index_replacement_string = ''.join(index_replacement_in_tuple)

        slide_order_to_replace = r's:11:"slide_order";s:1:"1"'
        slide_order_replacement_in_tuple = ('s:11:"slide_order";s:', str(len(str(key+1))), ':"', str(key+1), '"')
        slide_order_replacement_string = ''.join(slide_order_replacement_in_tuple)

        title_to_replace = r's:15:"Bozeman Science"'
        title_replacement_in_tuple = ('s:', str(len(video1[row][key+1][TITLE])), ':"', video1[row][key+1][TITLE], '"')
        title_replacement_string = ''.join(title_replacement_in_tuple)

        size_id_to_replace = r's:11:"wxvERNlUdBQ"'
        size_id_replacement_in_tuple = ('s:', str(len(video1[row][key+1][ID])), ':"', video1[row][key+1][ID], '"')
        size_id_replacement_string = ''.join(size_id_replacement_in_tuple)

        id_to_replace = r'wxvERNlUdBQ'
        id_replacement_string = video1[row][key+1][ID]

        start_time_to_replace = r's:4:"4:30"'
        start_time_replacement_in_tuple = ('s:', str(len(video1[row][key+1][START_TIME])), ':"', video1[row][key+1][START_TIME], '"')
        start_time_replacement_string = ''.join(start_time_replacement_in_tuple)

        end_time_to_replace = r's:4:"9:06"'
        end_time_replacement_in_tuple = ('s:', str(len(video1[row][key+1][END_TIME])), ':"', video1[row][key+1][END_TIME], '"')
        end_time_replacement_string = ''.join(end_time_replacement_in_tuple)

        sub_index = re.sub(index_to_replace, index_replacement_string, txtfile)
        sub_slide_order = re.sub(slide_order_to_replace, slide_order_replacement_string, sub_index)
        sub_title = re.sub(title_to_replace, title_replacement_string, sub_slide_order)
        sub_size_id = re.sub(size_id_to_replace, size_id_replacement_string, sub_title)
        sub_id = re.sub(id_to_replace, id_replacement_string, sub_size_id)
        sub_start_time = re.sub(start_time_to_replace, start_time_replacement_string, sub_id)
        sub_end_time = re.sub(end_time_to_replace, end_time_replacement_string, sub_start_time)

        self.list_body_text.append(sub_end_time)

    def modify_Body2(self, row, key, txtfile):
        index_to_replace = r'xi:5';
        index_replacement_in_tuple = ('i:', str(key))
        index_replacement_string = ''.join(index_replacement_in_tuple)

        slide_order_to_replace = r's:11:"slide_order";s:1:"6"'
        slide_order_replacement_in_tuple = ('s:11:"slide_order";s:', str(len(str(key+1))), ':"', str(key+1), '"')
        slide_order_replacement_string = ''.join(slide_order_replacement_in_tuple)

        title_to_replace = r's:11:"David Walz "'
        title_replacement_in_tuple = ('s:', str(len(video1[row][key+1][TITLE])), ':"', video1[row][key+1][TITLE], '"')
        title_replacement_string = ''.join(title_replacement_in_tuple)

        size_id_to_replace = r's:11:"gGsiCP44nms"'
        size_id_replacement_in_tuple = ('s:', str(len(video1[row][key+1][ID])), ':"', video1[row][key+1][ID], '"')
        size_id_replacement_string = ''.join(size_id_replacement_in_tuple)

        id_to_replace = r'gGsiCP44nms'
        id_replacement_string = video1[row][key+1][ID]

        start_time_to_replace = r's:0:"xstart"'
        start_time_replacement_in_tuple = ('s:', str(len(video1[row][key+1][START_TIME])), ':"', video1[row][key+1][START_TIME], '"')
        start_time_replacement_string = ''.join(start_time_replacement_in_tuple)

        end_time_to_replace = r's:0:"xend"'
        end_time_replacement_in_tuple = ('s:', str(len(video1[row][key+1][END_TIME])), ':"', video1[row][key+1][END_TIME], '"')
        end_time_replacement_string = ''.join(end_time_replacement_in_tuple)

        sub_index = re.sub(index_to_replace, index_replacement_string, txtfile)
        sub_slide_order = re.sub(slide_order_to_replace, slide_order_replacement_string, sub_index)
        sub_title = re.sub(title_to_replace, title_replacement_string, sub_slide_order)
        sub_size_id = re.sub(size_id_to_replace, size_id_replacement_string, sub_title)
        sub_id = re.sub(id_to_replace, id_replacement_string, sub_size_id)
        sub_start_time = re.sub(start_time_to_replace, start_time_replacement_string, sub_id)
        sub_end_time = re.sub(end_time_to_replace, end_time_replacement_string, sub_start_time)

        self.list_body_text.append(sub_end_time)

    #combine all the bodies
    #each body contain information for a single video
    def joinBody(self, row):
        self.modify_Body(row, self.body_text01, self.body_text02)
        return self.list_body_text
    
    def modify_SystemId(self, txtfile):
        base_system_id_file = open(txtfile, 'r+')
        base_system_id = int(base_system_id_file.read())
        list_i = 0
        for Id in range(base_system_id, base_system_id + len(self.list_body_text)):
            systemId_to_replace = r'a:5:{s:2:"id";s:4:"1962"'
            systemId_replacement_in_tuple = ('a:5:{s:2:"id";s:4:"', str(Id) ,'"')
            systemId_replacement_string = ''.join(systemId_replacement_in_tuple)
            sub_systemId = re.sub(systemId_to_replace, systemId_replacement_string, self.list_body_text[list_i])
            self.list_body_text[list_i] = sub_systemId
            list_i += 1


    def empty(self):
        self.list_body_text = []

    #export the txt which is read to be imported
    def export(self,row): 
        #modify the header of template
        self.modify_HeadSegment(row)
        self.modify_HeadSlidesN(row)

        #modify body
        body_in_list = self.joinBody(row)
        body = ''.join(body_in_list)
        self.modify_SystemId('systemId.txt')        
        self.body_text = body


        final = self.head_text + self.body_text + self.end_text0      #concatenate the header, the body and the footer
        test = open("test.txt",'w')
        test.truncate()
        test.write(final)



template = Output("template_head.txt","template_body1.txt","template_end.txt")
#template.export(2)
#template.modify_HeadSlidesN(0)
#template.modify_HeadSegment(0)
#print(template.head_text)
#template.modify_SystemId('systemID.txt')