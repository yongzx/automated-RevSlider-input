import openpyxl
import copy
import re
import os
import sys
import zipfile

# category --> title, id, start time, end time
TITLE, ID, START_TIME, END_TIME = range(4)


class ExcelData:
    """
    ExcelData contains all the data in a Excel sheet

    Arguments:
    None

    Container: list
    To access individual Topic data, ExcelData_instance[row_number_from_excel-2][column_number][category]

    Functions:
    read(TopicData object, Excel) : add all the information in the Excel into m_list. TopicData stores the information in a particular row.
    """

    def __init__(self):
        self.m_list = []

    def __getitem__(self, index):
        return self.m_list[index]

    def __len__(self):
        return len(self.m_list)

    def __str__(self):
        return str(self.m_list)

    def __repr__(self):
        return self.__str__()

    def read(self, topicdata, worksheet):
        """
        Append the Topicdata objects into the ExcelData _m.list 

        Arguments:
        self --- ExcelData
        topicdata --- TopicData object, which stores information of an individual cell in a dictionary container
        worksheet --- Excel worksheet

        Return:
        None

        """
        max_row = worksheet.max_row
        max_column = worksheet.max_column
        count_video = 1
        for row in range(2, max_row + 1):
            count_video = 1
            for col in range(4, max_column + 1):
                cell_var = worksheet.cell(
                    row=row, column=col).value  # cell_val is a string
                if cell_var is None or re.match('\S', cell_var) is None:
                    continue
                list_cell_var = cell_var.split('\n')
                # add the information in a cell into the Topicdata object
                topicdata.add(count_video, list_cell_var)
                count_video += 1

            # check if the row is empty ()
            if topicdata.isEmpty():
                continue

            # print(topicdata,'\n')
            topicdata.splitTime()  # split the duration into start time and end time
            # deepcopy to avoid aliasing
            topic_deepcopy = copy.deepcopy(topicdata)
            self.m_list.append(topic_deepcopy)
            topicdata.empty()  # empty the Topicdata object


class TopicData:
    """
    TopicData object contains the information in a particular row

    Arguments:
    None

    Container: dictionary
    key --> column number
    value --> list(   TITLE, (ID, (START_TIME, END_TIME))   ) 

    Functions:
    add() : add the video information of one cell (in list form) into the container. 
    empty() : empty the container
    split() : split the raw time input into start time and end time 
    """

    def __init__(self):
        self.m_dict = {}

    def __getitem__(self, key):
        return self.m_dict[key]

    def __len__(self):
        return len(self.m_dict)

    def __str__(self):
        return str(self.m_dict)

    def __repr__(self):
        return self.__str__()

    # implement deepcopy to avoid aliasing when the TopicData instance is
    # appended into VideoData instance
    def __deepcopy__(self, memo):
        return copy.deepcopy(self.m_dict)

    def add(self, index, list_var):
        """
        Append the information in a cell into the m_dict.

        Arguments: 
        self --- TopicData object
        index (key) --- indexing number, starts from 1
        list_var (value) --- a list of 3 things in each column: title, iD, raw time

        Return: 
        None
        """
        self.m_dict[index] = list_var

    def empty(self):
        """empty the m_dict """
        self.m_dict = {}

    def isEmpty(self):
        """return True if the m_dict is empty"""
        return not len(self.m_dict)

    def splitTime(self):
        """
        split the raw time into start time and end time

        return: None

        """
        for i in range(1, len(self.m_dict) + 1):

            split_time_list = []
            split_time_list.append(self.m_dict[i][TITLE])
            split_time_list.append(self.m_dict[i][ID])

            # start from 2 to ignore the title and the first ID
            for raw_time_input in range(2, len(self.m_dict[i])):

                temp_str = self.m_dict[i][raw_time_input]

                # if the raw time is 'whole' --> split into '' and ''
                if temp_str == 'whole':
                    split_time_list.append('')
                    split_time_list.append('')

                # split raw time into start time and end time
                else:
                    if self.isTime(temp_str) is False:
                        split_time_list.append(temp_str)  # put ID back

                    else:
                        temp_list = temp_str.split('-')
                        for j in range(0, len(temp_list)):
                            if temp_list[j] == 'start' or temp_list[j] == 'end':
                                temp_list[j] = ''
                        split_time_list.append(temp_list[0])
                        split_time_list.append(temp_list[1])
            self.m_dict[i] = split_time_list

    def isTime(self, string):
        """
        check if the string is ID or raw time

        Arguments:
        self --- TopicData
        string --- the information in the cell which has been stored in m_dict value

        Return:
        bool --- True if raw time, False if ID
        """
        temp_list = string.split('-')

        # ID may contain hyphen
        if len(temp_list) == 2:
            # raw time consists of actual time, start, or end
            if temp_list[0] == 'start' or temp_list[1] == 'end':
                return True
            elif re.search(":", string):
                return True
            else:
                return False
        else:
            # for case where ID doesn't contain hyphen
            return False


class Output:
    """
    Read template files and replace the information in the template file with the data from the ExcelData.

    Arguments:
    header template
    body template
    footer template

    Functions:
    modify_HeadSegment() : change title in header
    modify_HeadSlidesN() : change total number of slides in header
    modify_Body() : check the amount of information of a cell --> control flow for modify_Body1() and modify_Body2()
    modify_Body1() : modify the body template in the cell and put it in a list to be concatenated later (for cell information == 4)
    modify_Body2() : modify the body template in the cell and put it in a list to be concatenated later (for cell information > 4)
    joinBody(): call modify_Body()
    modify_SystemId(): Read in a unique ID and modify the body template
    empty(): resets the container holding the modified body templates, index, and slide order
    makedir(): create a folder
    cddir(): enter a folder
    exitdir(): exit from a folder
    export(): export multiple .txt files, which are ready to be imported into Wordpress database, into a newly created folder  


    """

    def __init__(self, head_file, body_file, end_file):
        tmp_head_file = open(os.path.join(os.path.dirname(
            os.path.abspath(__file__)), "Templates", head_file), 'r+')
        tmp_body_file = open(os.path.join(os.path.dirname(
            os.path.abspath(__file__)), "Templates", body_file), 'r+')
        tmp_end_file = open(os.path.join(os.path.dirname(
            os.path.abspath(__file__)), "Templates", end_file), 'r+')

        self.head_text0 = tmp_head_file.read()
        self.body_text0 = tmp_body_file.read()
        self.end_text0 = tmp_end_file.read()

        self.head_text = copy.deepcopy(self.head_text0)
        self.body_text = copy.deepcopy(self.body_text0)

        self.list_body_text = []
        self.SLIDE_ORDER = 1
        self.INDEX = 0

    def modify_HeadSegment(self, row):
        """
        modify the title for the to-be-imported file.
        """
        string_to_replace = r's:5:"title";s:2:"i1";s:5:"alias";s:2:"i1";s:9:"shortcode";s:23:"\[rev_slider alias="i1"]"'
        replacement_in_tuple = ('s:5:"title";s:2:"i', str(row), '";s:5:"alias";s:2:"i', str(row),
                                '";s:9:"shortcode";s:23:"[rev_slider alias="i', str(row), '"]"')

        replacement_string = ''.join(replacement_in_tuple)
        changed = re.sub(string_to_replace,
                         replacement_string, self.head_text0)
        self.head_text = changed

    def modify_HeadSlidesN(self):
        """
        modify the total number of slides
        """
        string_to_replace = r's:17:"custom_javascript";s:0:"";}s:6:"slides";a:7:{'
        replacement_in_tuple = (
            's:17:"custom_javascript";s:0:"";}s:6:"slides";a:', str(self.INDEX), ':{')
        replacement_string = ''.join(replacement_in_tuple)
        changed = re.sub(string_to_replace, replacement_string, self.head_text)
        self.head_text = changed

    def modify_Body(self, row, body_text, video_class):
        """
        check the amount of information in a particular cell

        Arguments:
        row --- particular row in an excel sheet. accessed from Exceldata object.
        body_text --- body template .txt file
        video_class --- Exceldata object
        """
        for key in range(1, len(video_class[row]) + 1):
            if len(video_class[row][key]) == 4:
                self.modify_Body1(row, key, body_text,
                                  video_class)  # create a slide
            elif len(video_class[row][key]) > 4:
                # create multiple slides
                self.modify_Body2(row, key, body_text, video_class)

    def modify_Body1(self, row, key, txtfile, video_class):
        """
        modify the body template in the cell and put it in a list to be concatenated later.
        (for cell information == 4)

        Arguments: 
        row --- particular row in an excel sheet. accessed from Exceldata object.
        key --- key-value pair, in which the key is paired with the information in a particular cell.  
        txtfile --- body template .txt file
        video_class --- Exceldata object

        """

        # index
        index_to_replace = r'xi:0'
        index_replacement_in_tuple = ('i:', str(self.INDEX))
        index_replacement_string = ''.join(index_replacement_in_tuple)
        self.INDEX += 1

        # slide order
        slide_order_to_replace = r's:11:"slide_order";s:1:"1"'
        slide_order_replacement_in_tuple = ('s:11:"slide_order";s:', str(
            len(str(self.SLIDE_ORDER))), ':"', str(self.SLIDE_ORDER), '"')
        slide_order_replacement_string = ''.join(
            slide_order_replacement_in_tuple)
        self.SLIDE_ORDER += 1

        # title of the video
        title_to_replace = r's:15:"Bozeman Science"'
        title_replacement_in_tuple = ('s:', str(
            len(video_class[row][key][TITLE])), ':"', video_class[row][key][TITLE], '"')
        title_replacement_string = ''.join(title_replacement_in_tuple)

        # size of ID
        size_id_to_replace = r's:11:"wxvERNlUdBQ"'
        size_id_replacement_in_tuple = ('s:', str(
            len(video_class[row][key][ID])), ':"', video_class[row][key][ID], '"')
        size_id_replacement_string = ''.join(size_id_replacement_in_tuple)

        # ID
        id_to_replace = r'wxvERNlUdBQ'
        id_replacement_string = video_class[row][key][ID]

        # starting time
        start_time_to_replace = r's:4:"4:30"'
        start_time_replacement_in_tuple = ('s:', str(len(video_class[row][key][
                                           START_TIME])), ':"', video_class[row][key][START_TIME], '"')
        start_time_replacement_string = ''.join(
            start_time_replacement_in_tuple)

        # ending time
        end_time_to_replace = r's:4:"9:06"'
        end_time_replacement_in_tuple = ('s:', str(
            len(video_class[row][key][END_TIME])), ':"', video_class[row][key][END_TIME], '"')
        end_time_replacement_string = ''.join(end_time_replacement_in_tuple)

        # modifying process
        sub_index = re.sub(index_to_replace, index_replacement_string, txtfile)
        sub_slide_order = re.sub(
            slide_order_to_replace, slide_order_replacement_string, sub_index)
        sub_title = re.sub(
            title_to_replace, title_replacement_string, sub_slide_order)
        sub_size_id = re.sub(size_id_to_replace,
                             size_id_replacement_string, sub_title)
        sub_id = re.sub(id_to_replace, id_replacement_string, sub_size_id)
        sub_start_time = re.sub(start_time_to_replace,
                                start_time_replacement_string, sub_id)
        sub_end_time = re.sub(end_time_to_replace,
                              end_time_replacement_string, sub_start_time)

        self.list_body_text.append(sub_end_time)

    def modify_Body2(self, row, key, txtfile, video_class):
        """
        modify the body template in the cell and put it in a list to be concatenated later
        (for cell information > 4)

        Arguments: 
        row --- particular row in an excel sheet. accessed from Exceldata object.
        key --- key-value pair, in which the key is paired with the information in a particular cell.  
        txtfile --- body template .txt file
        video_class --- Exceldata object
        """

        loop_index = 1  # negate the name
        title_index = 1
        total_title_index = 0

        while (loop_index < len(video_class[row][key])):
            total_title_index += 1

            if len(video_class[row][key][loop_index]) > 6:
                loop_index += 1

            if len(video_class[row][key][loop_index]) <= 6:
                loop_index += 1
                loop_index += 1

        loop_index = 1  # reset loop index

        while (loop_index < len(video_class[row][key])):
            title_pre_replacement = video_class[row][key][
                TITLE] + ' ' + str(title_index) + '/' + str(total_title_index)
            title_index += 1

            if len(video_class[row][key][loop_index]) > 6:
                id_pre_replacement = str(video_class[row][key][loop_index])
                loop_index += 1

            if len(video_class[row][key][loop_index]) <= 6:
                start_time_pre_replacement = str(
                    video_class[row][key][loop_index])
                loop_index += 1
                end_time_pre_replacement = str(
                    video_class[row][key][loop_index])
                loop_index += 1

            index_to_replace = r'xi:0'
            index_replacement_in_tuple = ('i:', str(self.INDEX))
            index_replacement_string = ''.join(index_replacement_in_tuple)
            self.INDEX += 1

            slide_order_to_replace = r's:11:"slide_order";s:1:"1"'
            slide_order_replacement_in_tuple = ('s:11:"slide_order";s:', str(
                len(str(self.SLIDE_ORDER))), ':"', str(self.SLIDE_ORDER), '"')
            slide_order_replacement_string = ''.join(
                slide_order_replacement_in_tuple)
            self.SLIDE_ORDER += 1

            title_to_replace = r's:15:"Bozeman Science"'
            title_replacement_in_tuple = (
                's:', str(len(title_pre_replacement)), ':"', title_pre_replacement, '"')
            title_replacement_string = ''.join(title_replacement_in_tuple)

            size_id_to_replace = r's:11:"wxvERNlUdBQ"'
            size_id_replacement_in_tuple = (
                's:', str(len(id_pre_replacement)), ':"', id_pre_replacement, '"')
            size_id_replacement_string = ''.join(size_id_replacement_in_tuple)

            id_to_replace = r'wxvERNlUdBQ'
            id_replacement_string = id_pre_replacement

            start_time_to_replace = r's:4:"4:30"'
            start_time_replacement_in_tuple = ('s:', str(
                len(start_time_pre_replacement)), ':"', start_time_pre_replacement, '"')
            start_time_replacement_string = ''.join(
                start_time_replacement_in_tuple)

            end_time_to_replace = r's:4:"9:06"'
            end_time_replacement_in_tuple = (
                's:', str(len(end_time_pre_replacement)), ':"', end_time_pre_replacement, '"')
            end_time_replacement_string = ''.join(
                end_time_replacement_in_tuple)

            sub_index = re.sub(
                index_to_replace, index_replacement_string, txtfile)
            sub_slide_order = re.sub(
                slide_order_to_replace, slide_order_replacement_string, sub_index)
            sub_title = re.sub(
                title_to_replace, title_replacement_string, sub_slide_order)
            sub_size_id = re.sub(size_id_to_replace,
                                 size_id_replacement_string, sub_title)
            sub_id = re.sub(id_to_replace, id_replacement_string, sub_size_id)
            sub_start_time = re.sub(
                start_time_to_replace, start_time_replacement_string, sub_id)
            sub_end_time = re.sub(end_time_to_replace,
                                  end_time_replacement_string, sub_start_time)

            self.list_body_text.append(sub_end_time)

    # combine all the bodies
    # each body represents individual cell
    def joinBody(self, row, video_class):
        """call the modify_Body"""
        self.modify_Body(row, self.body_text0, video_class)
        return self.list_body_text

    def modify_SystemId(self, txtfile):
        """
        Read in a unique ID. 
        Since currently wordpress database automatically updates the ID of the imported file, this member function
            does not update the system ID in the txtfile from which it reads in the system ID. 

        Arguments:
        txtfile --- contains the ID number

        Returns:
        None

        """
        base_system_id_file = open(os.path.join(os.path.dirname(
            os.path.abspath(__file__)), "Templates", txtfile), 'r+')
        base_system_id = int(base_system_id_file.read())
        list_i = 0
        for Id in range(base_system_id, base_system_id + len(self.list_body_text)):
            systemId_to_replace = r'a:5:{s:2:"id";s:4:"1962"'
            systemId_replacement_in_tuple = (
                'a:5:{s:2:"id";s:4:"', str(Id), '"')
            systemId_replacement_string = ''.join(
                systemId_replacement_in_tuple)
            sub_systemId = re.sub(
                systemId_to_replace, systemId_replacement_string, self.list_body_text[list_i])
            self.list_body_text[list_i] = sub_systemId
            list_i += 1

    def empty(self):
        """resets:
        list_body_test --- which holds all the information to replace the body template file
        slide_order --- starts from 1
        index --- starts from 0
        """
        self.list_body_text = []
        self.SLIDE_ORDER = 1
        self.INDEX = 0

    def makedir(self, file_name):
        """create a folder"""
        try:
            if not os.path.isdir("../{0}".format(file_name)):
                os.makedirs(file_name)
        except FileExistsError:
            pass

    def cddir(self, file_name):
        """enter the folder"""
        try:
            os.chdir(file_name)
        except:
            print(
                "Unable to find the folder.\nPlease create a folder manually and try again.")
            print("Folder name: Exported excel_name.xlsx sheet_name")
            sys.exit()

    def exitdir(self):
        """exit from the folder to the main folder automated_RevSlider_input"""
        os.chdir("..")

    # export the txt which is read to be imported
    def export(self, excel_name, sheet_name, video_class):
        """
        Create a folder and save multiple the txt files which are ready to be exported into the folder

        Arguments:
        excel_name --- name of the excel file
        sheet_name --- name of the sheet
        video_class --- ExcelData which contains all the information contained in the sheet.

        Returns:
        .txt files
        """
        file_name = "Exported {0} {1}".format(excel_name, sheet_name)
        
        self.makedir(file_name)

        for num_slides in range(1, len(video_class) + 1):
            # modify body
            # minus one because the index of m_list of ExcelData starts from 0
            body_in_list = self.joinBody(num_slides - 1, video_class)
            body = ''.join(body_in_list)
            self.modify_SystemId('systemId.txt')
            self.body_text = body

            # modify the header of template
            # placed after body is modified because updated INDEX is needed
            self.modify_HeadSegment(num_slides)
            self.modify_HeadSlidesN()

            # concatenate the header, the body and the footer
            final = self.head_text + self.body_text + self.end_text0

            self.cddir(file_name)
            file = FileOutput()
            file.create_txt(final)
            file.create_zip(num_slides)
            file.remove()
            self.exitdir()

            self.empty()

class FileOutput:

    def __init__(self):
        self.text = "slider_export.txt"

    def create_txt(self, content):
        with open(self.text, "w+") as f:
            f.truncate()
            f.write(content)
            print("Generated: {0}".format(self.text))

    def create_zip(self, num_slides):
        zf = zipfile.ZipFile('i{0}.zip'.format(num_slides), 'w')
        try:
            print('Zip {0} into i{1}.zip\n'.format(self.text,num_slides))
            zf.write(self.text)
        finally:
            zf.close()

    def remove(self):
        os.remove(self.text)

class FileNameError(Exception):
    pass


def readFile():
    """
    Prompt for the name of the excel file.
    Read in and open the excel file.

    Arguments:
    None

    Returns:
    name of the excel file --- str
    workbook ---openpyxl
    """
    while True:
        try:
            excel_name = input(
                "Name of your Excel document (remember to put .xlsx in your input, e.g., test.xlsx): ")
            if re.search(".xlsx", excel_name) is None:
                raise FileNameError
            excel_file = openpyxl.load_workbook(excel_name)
            return excel_name, excel_file
        except FileNameError:
            print("You forgot to put .xlsx in your file name. \n")
        except FileNotFoundError:
            print("File '{0}' not found. Remember to put your file in the same folder as this script.py. \n".format(
                excel_name))


def readSheet(excel_file):
    """
    Prompt for the name of sheet
    Read in and open a particular sheet in the excel file

    Arguments: 
    excel_file ---openpyxl workbook

    Return:
    name of the sheet --- str
    sheet --- openpyxl

    """
    while True:
        try:
            sheet_name = input("Sheet name (case sensitive): ")
            sheet = excel_file.get_sheet_by_name(sheet_name)
            return sheet_name, sheet
        except KeyError:
            print("Sheet '{0}' does not exist. \n".format(sheet_name))


if __name__ == "__main__":
    video = ExcelData()
    topic = TopicData()
    template = Output("template_head.txt",
                      "template_body1.txt", "template_end.txt")

    # obtain excel file and sheet name
    excel_name, excel_file = readFile()
    sheet_name, sheet = readSheet(excel_file)

    # processing information in the particular sheet
    video.read(topic, sheet)

    # output
    print("Total slides: ", len(video))
    template.export(excel_name, sheet_name, video)
