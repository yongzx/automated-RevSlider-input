import openpyxl
import copy
import re
class ExcelData:
    """
    ExcelData contains all the data in a Excel sheet

    Arguments:
    None

    Container: list
    To access individual Topic data, ExcelData_instance[row_number_from_excel-2][column_number][category]

    Functions:
    read(TopicData object, Excel) : add all the information in the Excel into m_list. TopicData stores the information in a particular row.
    """

    def __init__(self):
        self.m_list = []

    def __getitem__(self, index):
        return self.m_list[index]

    def __len__(self):
        return len(self.m_list)

    def __str__(self):
        return str(self.m_list)

    def __repr__(self):
        return self.__str__()

    def read(self, topicdata, worksheet):
        """
        Append the Topicdata objects into the ExcelData _m.list 

        Arguments:
        self --- ExcelData
        topicdata --- TopicData object, which stores information of an individual cell in a dictionary container
        worksheet --- Excel worksheet

        Return:
        None

        """
        max_row = worksheet.max_row
        max_column = worksheet.max_column
        count_video = 1
        for row in range(2, max_row + 1):
            count_video = 1
            for col in range(4, max_column + 1):
                cell_var = worksheet.cell(
                    row=row, column=col).value  # cell_val is a string
                if cell_var is None or re.match('\S', cell_var) is None:
                    continue
                list_cell_var = cell_var.split('\n')
                # add the information in a cell into the Topicdata object
                topicdata.add(count_video, list_cell_var)
                count_video += 1

            # check if the row is empty ()
            if topicdata.isEmpty():
                continue

            # print(topicdata,'\n')
            topicdata.splitTime()  # split the duration into start time and end time
            # deepcopy to avoid aliasing
            topic_deepcopy = copy.deepcopy(topicdata)
            self.m_list.append(topic_deepcopy)
            topicdata.empty()  # empty the Topicdata object